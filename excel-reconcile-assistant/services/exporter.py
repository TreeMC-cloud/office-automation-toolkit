"""导出工作簿 — 中文 Sheet 名、表头样式、列宽自适应、冻结首行、差异高亮"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# 样式常量
_HEADER_FONT = Font(name="Microsoft YaHei UI", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="center", wrap_text=False)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_DIFF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_DIFF_FONT = Font(color="9C0006")
_ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def build_export_workbook(
    results: dict,
    duplicates_a: pd.DataFrame,
    duplicates_b: pd.DataFrame,
    fuzzy_matches: pd.DataFrame,
    report_text: str,
) -> bytes:
    """构建格式化的 Excel 工作簿，返回 bytes"""
    wb = Workbook()

    # 核对报告（第一个 Sheet）
    ws_report = wb.active
    ws_report.title = "核对报告"
    _write_report_sheet(ws_report, report_text)

    # 数据 Sheet
    sheet_defs = [
        ("匹配结果", results.get("matched_records", pd.DataFrame())),
        ("完全匹配", results.get("exact_matches", pd.DataFrame())),
        ("A缺失于B", results.get("missing_in_b", pd.DataFrame())),
        ("B缺失于A", results.get("missing_in_a", pd.DataFrame())),
        ("差异明细", results.get("mismatch_details", pd.DataFrame())),
        ("重复记录A", duplicates_a),
        ("重复记录B", duplicates_b),
        ("模糊匹配", fuzzy_matches),
    ]

    for sheet_name, df in sheet_defs:
        if df is None or df.empty:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["（无数据）"])
            continue
        ws = wb.create_sheet(title=sheet_name)
        _write_data_sheet(ws, df, highlight_diff=(sheet_name == "差异明细"))

    # 摘要 Sheet
    ws_summary = wb.create_sheet(title="统计摘要")
    _write_summary_sheet(ws_summary, results.get("stats", {}))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_report_sheet(ws, report_text: str) -> None:
    """写入核对报告文本"""
    ws.column_dimensions["A"].width = 80
    for i, line in enumerate(report_text.split("\n"), 1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name="Microsoft YaHei UI", size=11)
        cell.alignment = Alignment(vertical="center")


def _write_data_sheet(ws, df: pd.DataFrame, highlight_diff: bool = False) -> None:
    """写入数据表，带格式化"""
    columns = df.columns.tolist()

    # 表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # 数据行
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_alt = row_idx % 2 == 0
        for col_idx, col_name in enumerate(columns, 1):
            value = row[col_name]
            # NaN → 空字符串显示
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER

            # 隔行变色
            if is_alt:
                cell.fill = _ALT_ROW_FILL

            # 差异明细高亮
            if highlight_diff and col_name in ("A值", "B值"):
                cell.fill = _DIFF_FILL
                cell.font = _DIFF_FONT

    # 冻结首行
    ws.freeze_panes = "A2"

    # 列宽自适应
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 102)):  # 采样前 100 行
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        # 中文字符宽度约为英文的 1.5 倍
        adjusted_width = min(max_len * 1.3 + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 自动筛选
    if columns:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(df) + 1}"


def _write_summary_sheet(ws, stats: dict) -> None:
    """写入统计摘要"""
    label_map = {
        "left_rows": "A 表记录数",
        "right_rows": "B 表记录数",
        "matched_rows": "成功匹配",
        "exact_match_rows": "完全一致",
        "mismatch_rows": "字段不一致",
        "missing_in_b_rows": "A 有 B 无",
        "missing_in_a_rows": "B 有 A 无",
        "match_rate": "匹配率",
        "exact_rate": "一致率",
    }

    # 表头
    for col_idx, header in enumerate(["指标", "数值"], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    row = 2
    for key, label in label_map.items():
        value = stats.get(key, "")
        ws.cell(row=row, column=1, value=label).border = _THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = _THIN_BORDER
        row += 1

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
