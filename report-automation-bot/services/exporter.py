"""Excel 导出 — 中文 Sheet、表头样式、列宽自适应、冻结首行、图表嵌入"""

from __future__ import annotations

import base64
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(name="Microsoft YaHei UI", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def build_report_workbook(
    raw_dataframe: pd.DataFrame,
    overview: dict,
    trend_dataframe: pd.DataFrame,
    dimension_dataframe: pd.DataFrame,
    summary_text: str,
    trend_chart_b64: str = "",
    dimension_chart_b64: str = "",
) -> bytes:
    wb = Workbook()

    # 摘要
    ws_summary = wb.active
    ws_summary.title = "分析摘要"
    _write_summary(ws_summary, overview, summary_text)

    # 趋势数据
    ws_trend = wb.create_sheet("趋势数据")
    _write_data_sheet(ws_trend, trend_dataframe)
    if trend_chart_b64:
        _embed_chart(ws_trend, trend_chart_b64, len(trend_dataframe) + 3)

    # 维度数据
    ws_dim = wb.create_sheet("维度分析")
    _write_data_sheet(ws_dim, dimension_dataframe)
    if dimension_chart_b64:
        _embed_chart(ws_dim, dimension_chart_b64, len(dimension_dataframe) + 3)

    # 原始数据
    ws_raw = wb.create_sheet("原始数据")
    _write_data_sheet(ws_raw, raw_dataframe)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary(ws, overview: dict, summary_text: str) -> None:
    """写入摘要 Sheet"""
    label_map = {
        "total_records": "总记录数",
        "total_value": "总指标值",
        "average_value": "均值",
        "median_value": "中位数",
        "max_value": "最大值",
        "min_value": "最小值",
        "std_value": "标准差",
        "latest_period": "最新周期",
        "latest_period_value": "最新周期值",
    }

    # 表头
    for col_idx, header in enumerate(["指标", "数值"], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    row = 2
    for key, label in label_map.items():
        val = overview.get(key, "")
        ws.cell(row=row, column=1, value=label).border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=val)
        c.border = _THIN_BORDER
        if isinstance(val, float):
            c.number_format = "#,##0.00"
        row += 1

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20

    # 摘要文本
    row += 1
    ws.cell(row=row, column=1, value="分析摘要").font = Font(bold=True, size=12)
    row += 1
    for line in summary_text.split("\n"):
        ws.cell(row=row, column=1, value=line)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1


def _write_data_sheet(ws, df: pd.DataFrame) -> None:
    """写入数据表"""
    if df is None or df.empty:
        ws.append(["（无数据）"])
        return

    columns = df.columns.tolist()

    # 表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # 数据
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_alt = row_idx % 2 == 0
        for col_idx, col_name in enumerate(columns, 1):
            val = row[col_name]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER
            if is_alt:
                cell.fill = _ALT_FILL
            if isinstance(val, float):
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"

    # 列宽自适应
    for col_idx, col_name in enumerate(columns, 1):
        max_len = _display_width(str(col_name))
        for r in range(2, min(len(df) + 2, 102)):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, min(_display_width(str(v)), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    # 自动筛选
    if columns:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(df) + 1}"


def _embed_chart(ws, b64_str: str, start_row: int) -> None:
    """嵌入 base64 图表到 Excel"""
    try:
        img_data = base64.b64decode(b64_str)
        img = XlImage(BytesIO(img_data))
        img.width = 700
        img.height = 350
        ws.add_image(img, f"A{start_row}")
    except Exception as e:
        print(f"[exporter] 图表嵌入失败: {e}")


def _display_width(text: str) -> int:
    """估算文本显示宽度（中文字符算2，其他算1）"""
    width = 0
    for ch in text:
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
            width += 2
        else:
            width += 1
    return width
